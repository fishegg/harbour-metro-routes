import openpyxl

# book_file=input("input xlsx file name:")
book_file="D:\\project\\untitled1-master\\testdata\\新建 Microsoft Excel 工作表.xlsx"
# book_file="D:\\project\\untitled1-master\\testdata\\工作簿1.xlsx"
wb=openpyxl.load_workbook(book_file)
sheet=wb["Sheet1"]
# sheet=wb["Sheet2"]
# sheet=wb["Sheet4"]
while(True):
	option=int(input("1=fill with 0,2=fill line grids,3=fill interchange grids,4=fill with 0 by row and column,-1=quit:"))
	if(option != 1 and option != 2 and option != 3 and option != 4):
		exit()
	strings=[]
	#book_file="D:\\project\\ocrtest\\xlsx.xlsx"

	while(True):
		if(option==1):
			count=int(input("stations count:"))
			if(count == -1):
				break
			for row in range (1,count+1):
				for column in range (1,count+1):
					sheet.cell(row,column,0)
			wb.save(book_file)
		elif(option==2):
			terminal1=int(input("terminal 1:"))
			if(terminal1==-1):
				break
			terminal2=int(input("terminal 2:"))
			if(terminal2==-1):
				break
			distance=int(input("distance:"))
			if(distance==-1):
				break
			for row in range (terminal1,terminal2+1):
				for column in range (terminal1,terminal2+1):
					if(row==column+1 or column==row+1):
						sheet.cell(row,column,distance)
			wb.save(book_file)
		elif(option==3):
			interchange1=int(input("interchange 1:"))
			if(interchange1==-1):
				break
			interchange2=int(input("interchange 2:"))
			if(interchange2==-1):
				break
			distance=int(input("distance:"))
			if(distance==-1):
				break
			sheet.cell(interchange1,interchange2,distance)
			sheet.cell(interchange2,interchange1,distance)
			wb.save(book_file)
		elif(option==4):
			row_start=int(input("row start:"))
			if(row_start == -1):
				break
			row_end=int(input("row end:"))
			if(row_end == -1):
				break
			for row in range (row_start,row_end+1):
				for column in range (row_start,row_end+1):
					sheet.cell(row,column,0)
					sheet.cell(column,row,0)
			for row in range (row_start,row_end+1):
				for column in range (1,row_end):
					sheet.cell(row,column,0)
					sheet.cell(column,row,0)
			wb.save(book_file)
	wb.save(book_file)